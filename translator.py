import openpyxl as xl
from googletrans import Translator


file_name = "###.xlsx" # The name of your Excel file
fhandle = xl.load_workbook(f"{file_name}")
sheet = fhandle.active
translator = Translator()


def translate_line(text):
    translated = translator.translate(text, src="ja", dest="en").text
    return translated


workbook = xl.load_workbook(file_name)
active_sheet = workbook.active

for row_index in range(1, active_sheet.max_row + 1):
    cell_value = active_sheet.cell(row=row_index, column=1).value

    if cell_value:
        text_to_translate = cell_value
        translated_text = translate_line(text_to_translate)
        active_sheet.cell(row=row_index, column=2, value=translated_text)


workbook.save(f"Translated {file_name}")
workbook.close()
